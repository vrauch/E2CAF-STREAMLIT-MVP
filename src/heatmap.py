"""Maturity heatmap rendering and Excel export for completed assessments.

Given a list of domain scores (from Step 5), produces:
  - render_heatmap_html()  → Bootstrap HTML table for st.components.v1.html()
  - generate_heatmap_excel() → .xlsx bytes for st.download_button()

Per-level score formula:
    level_score(avg, L) = max(0, min(1, avg - (L - 1)))
This maps a 1–5 avg_score to a staircase of 0–1 completions per level.
"""

from __future__ import annotations

import io
from typing import Optional

# ── Domain brand colours (matches the HPE Excel template for the first 8,
#    extended for the 4 newer TMM domains) ──────────────────────────────────
DOMAIN_COLORS: dict[str, str] = {
    "Strategy & Governance":                  "#32DAC8",
    "Security":                               "#0D5265",
    "People":                                 "#7630EA",
    "Applications":                           "#BFBFBF",
    "Data":                                   "#FFC000",
    "DevOps":                                 "#01A982",
    "Innovation":                             "#FF8300",
    "Operations":                             "#CE67FF",
    # Newer domains
    "AI & Cognitive Systems":                 "#0066CC",
    "Intelligent Automation & Operations":    "#D04830",
    "Sustainability & Responsible Technology":"#059669",
    "Experience & Ecosystem Enablement":      "#9333EA",
}
_DEFAULT_COLOR = "#8A929A"

# ── Maturity level definitions (L5 → L1, matching app's DB naming) ─────────
LEVELS = [
    (5, "Adaptive"),
    (4, "Intelligent"),
    (3, "Integrated"),
    (2, "Defined"),
    (1, "Ad Hoc"),
]

# ── Cell colour thresholds ───────────────────────────────────────────────────
_GREEN  = "#B7E2CD"   # fully achieved  (score == 1.0)
_AMBER  = "#FDE9B2"   # partial         (0 < score < 1.0)
_WHITE  = "#FFFFFF"   # not yet reached (score == 0.0)
_GRAY   = "#D9D9D9"   # average row background


def _level_scores(avg_score: Optional[float]) -> list[float]:
    """Convert a 1–5 avg_score to 5 per-level 0–1 completion values (L1…L5)."""
    if avg_score is None:
        return [0.0] * 5
    s = float(avg_score)
    return [max(0.0, min(1.0, s - (lv - 1))) for lv in range(1, 6)]


def _cell_bg(score: float) -> str:
    if score >= 1.0:
        return _GREEN
    if score > 0.0:
        return _AMBER
    return _WHITE


def _text_color(bg: str) -> str:
    """Return dark or light text for readability on the given hex background."""
    # Security (#0D5265) and a few others are very dark — use white text
    dark_backgrounds = {"#0D5265", "#7630EA", "#9333EA", "#0066CC", "#D04830"}
    return "#FFFFFF" if bg in dark_backgrounds else "#1A1A2E"


# ────────────────────────────────────────────────────────────────────────────
# In-app HTML visualization
# ────────────────────────────────────────────────────────────────────────────

def render_heatmap_html(dom_scores: list[dict]) -> str:
    """
    Build a Bootstrap 5 HTML table heatmap from domain scores.

    dom_scores: list of dicts with keys: domain (str), avg_score (float),
                target (int), gap (float)
    """
    # Build lookup: domain → (avg_score, target)
    score_map: dict[str, dict] = {
        row["domain"]: row for row in dom_scores
    }
    domains = [row["domain"] for row in dom_scores]

    # ── Header cells ─────────────────────────────────────────────────────────
    header_cells = ""
    for d in domains:
        color = DOMAIN_COLORS.get(d, _DEFAULT_COLOR)
        tc    = _text_color(color)
        # Shorten very long domain names to two lines
        label = d.replace(" & ", "<br>&nbsp;")
        header_cells += (
            f'<th style="background:{color};color:{tc};text-align:center;'
            f'font-size:.72rem;font-weight:700;padding:.45rem .5rem;'
            f'min-width:90px;max-width:110px;vertical-align:middle;'
            f'word-break:break-word;">{label}</th>'
        )

    # ── Level rows (L5 → L1) ────────────────────────────────────────────────
    level_rows = ""
    for lv_num, lv_name in LEVELS:
        cells = ""
        for d in domains:
            row = score_map.get(d)
            if row is None:
                score = 0.0
            else:
                scores_list = _level_scores(row.get("avg_score"))
                score = scores_list[lv_num - 1]   # index 0 = L1
            bg = _cell_bg(score)
            pct = f"{score:.0%}" if score > 0 else "—"
            cells += (
                f'<td style="background:{bg};text-align:center;'
                f'font-family:\'JetBrains Mono\',monospace;font-size:.78rem;'
                f'font-weight:600;color:#1A1A2E;padding:.4rem .3rem;">{pct}</td>'
            )
        level_rows += (
            f'<tr>'
            f'<td style="background:#666666;color:#FFFFFF;font-weight:700;'
            f'font-size:.72rem;padding:.4rem .6rem;white-space:nowrap;">'
            f'L{lv_num}<br><span style="font-weight:400;font-size:.65rem;">{lv_name}</span>'
            f'</td>'
            f'{cells}</tr>\n'
        )

    # ── Average Maturity row ─────────────────────────────────────────────────
    avg_cells = ""
    for d in domains:
        row = score_map.get(d)
        val = f"{row['avg_score']:.1f}" if row and row.get("avg_score") is not None else "—"
        avg_cells += (
            f'<td style="background:{_GRAY};text-align:center;'
            f'font-family:\'JetBrains Mono\',monospace;font-size:.78rem;'
            f'font-weight:700;color:#1A1A2E;padding:.4rem .3rem;">{val}</td>'
        )

    # ── Target row ───────────────────────────────────────────────────────────
    tgt_cells = ""
    for d in domains:
        row  = score_map.get(d)
        val  = str(int(row["target"])) if row and row.get("target") is not None else "—"
        tgt_cells += (
            f'<td style="background:#EEF1F5;text-align:center;'
            f'font-family:\'JetBrains Mono\',monospace;font-size:.78rem;'
            f'font-weight:600;color:#5A6570;padding:.4rem .3rem;">{val}</td>'
        )

    # ── Legend ───────────────────────────────────────────────────────────────
    legend = (
        '<div style="display:flex;gap:1.2rem;margin-bottom:1rem;align-items:center;'
        'font-size:.72rem;color:#5A6570;">'
        f'<span style="display:flex;align-items:center;gap:.4rem;">'
        f'<span style="width:14px;height:14px;background:{_GREEN};border:1px solid #ccc;border-radius:2px;display:inline-block"></span>Fully achieved (100%)</span>'
        f'<span style="display:flex;align-items:center;gap:.4rem;">'
        f'<span style="width:14px;height:14px;background:{_AMBER};border:1px solid #ccc;border-radius:2px;display:inline-block"></span>Partial progress</span>'
        f'<span style="display:flex;align-items:center;gap:.4rem;">'
        f'<span style="width:14px;height:14px;background:{_WHITE};border:1px solid #ccc;border-radius:2px;display:inline-block"></span>Not yet reached</span>'
        '</div>'
    )

    return f"""<!DOCTYPE html>
<html><head>
<meta charset="UTF-8">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
<style>
  body {{ font-family:'Inter',sans-serif; background:#fff; padding:.5rem 0; }}
  table {{ border-collapse:collapse; width:100%; }}
  th, td {{ border:1px solid #E0E4EA; }}
</style>
</head><body>
{legend}
<div style="overflow-x:auto;">
<table>
<thead>
  <tr>
    <th style="background:#CCCCCC;font-size:.7rem;font-weight:700;
               padding:.4rem .6rem;text-transform:uppercase;letter-spacing:.06em;
               color:#333;white-space:nowrap;">Level</th>
    {header_cells}
  </tr>
</thead>
<tbody>
{level_rows}
  <tr>
    <td style="background:{_GRAY};font-size:.72rem;font-weight:700;
               padding:.4rem .6rem;color:#1A1A2E;white-space:nowrap;">Avg Maturity<br>
      <span style="font-weight:400;font-size:.65rem;color:#5A6570;">(1–5 scale)</span>
    </td>
    {avg_cells}
  </tr>
  <tr>
    <td style="background:#EEF1F5;font-size:.72rem;font-weight:700;
               padding:.4rem .6rem;color:#5A6570;white-space:nowrap;">Target</td>
    {tgt_cells}
  </tr>
</tbody>
</table>
</div>
</body></html>"""


# ────────────────────────────────────────────────────────────────────────────
# Excel export
# ────────────────────────────────────────────────────────────────────────────

def generate_heatmap_excel(
    dom_scores: list[dict],
    client_name: str = "",
    engagement_name: str = "",
    use_case_name: str = "",
) -> bytes:
    """
    Generate a .xlsx heatmap matching the HPE assessment-heatmap style.
    Returns raw bytes suitable for st.download_button().
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    def _hex_fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))

    def _thin_border() -> Border:
        s = Side(style="thin", color="E0E4EA")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()
    ws = wb.active
    ws.title = "Maturity Heatmap"

    domains = [row["domain"] for row in dom_scores]
    score_map = {row["domain"]: row for row in dom_scores}

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws["B1"] = "Maturity Heatmap"
    ws["B1"].font = Font(bold=True, size=14)
    ws["B1"].fill = _hex_fill("CCCCCC")

    # ── Row 2: Metadata ───────────────────────────────────────────────────────
    meta_parts = []
    if client_name:
        meta_parts.append(f"Client: {client_name}")
    if engagement_name:
        meta_parts.append(f"Engagement: {engagement_name}")
    if use_case_name:
        meta_parts.append(f"Use Case: {use_case_name}")
    ws["B2"] = "  |  ".join(meta_parts)
    ws["B2"].font = Font(size=9, color="5A6570")

    # ── Row 4: Domain headers ─────────────────────────────────────────────────
    ws["B4"] = "Level"
    ws["B4"].font      = Font(bold=True, size=9, color="333333")
    ws["B4"].fill      = _hex_fill("CCCCCC")
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B4"].border    = _thin_border()

    for col_idx, d in enumerate(domains, start=3):   # columns C onwards
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}4"]
        color = DOMAIN_COLORS.get(d, _DEFAULT_COLOR)
        tc    = "FFFFFF" if color in {
            "#0D5265", "#7630EA", "#9333EA", "#0066CC", "#D04830"
        } else "1A1A2E"
        cell.value     = d
        cell.font      = Font(bold=True, size=8, color=tc)
        cell.fill      = _hex_fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _thin_border()
        ws.column_dimensions[col_letter].width = 14

    ws.column_dimensions["B"].width = 16
    ws.row_dimensions[4].height     = 30

    # ── Rows 5–9: Level rows (L5 → L1) ───────────────────────────────────────
    for row_offset, (lv_num, lv_name) in enumerate(LEVELS):
        xls_row = 5 + row_offset
        # Level label cell
        label_cell = ws[f"B{xls_row}"]
        label_cell.value     = f"L{lv_num} — {lv_name}"
        label_cell.font      = Font(bold=True, size=8, color="FFFFFF")
        label_cell.fill      = _hex_fill("666666")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border    = _thin_border()
        ws.row_dimensions[xls_row].height = 20

        for col_idx, d in enumerate(domains, start=3):
            col_letter = get_column_letter(col_idx)
            row_data = score_map.get(d)
            if row_data is None:
                score = 0.0
            else:
                scores_list = _level_scores(row_data.get("avg_score"))
                score = scores_list[lv_num - 1]

            cell = ws[f"{col_letter}{xls_row}"]
            cell.value     = round(score, 2)
            cell.number_format = "0.00"
            cell.font      = Font(bold=True, size=9)
            cell.fill      = _hex_fill(_cell_bg(score))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _thin_border()

    # ── Row 10: Average Maturity ──────────────────────────────────────────────
    avg_row = 10
    avg_cell = ws[f"B{avg_row}"]
    avg_cell.value     = "Avg Maturity (1–5)"
    avg_cell.font      = Font(bold=True, size=8)
    avg_cell.fill      = _hex_fill("D9D9D9")
    avg_cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
    avg_cell.border    = _thin_border()
    ws.row_dimensions[avg_row].height = 22

    for col_idx, d in enumerate(domains, start=3):
        col_letter = get_column_letter(col_idx)
        row_data = score_map.get(d)
        val = round(float(row_data["avg_score"]), 1) if row_data and row_data.get("avg_score") is not None else None
        cell = ws[f"{col_letter}{avg_row}"]
        cell.value     = val
        cell.number_format = "0.0"
        cell.font      = Font(bold=True, size=9)
        cell.fill      = _hex_fill("D9D9D9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

    # ── Row 11: Target Maturity ───────────────────────────────────────────────
    tgt_row = 11
    tgt_cell = ws[f"B{tgt_row}"]
    tgt_cell.value     = "Target Maturity"
    tgt_cell.font      = Font(bold=True, size=8)
    tgt_cell.fill      = _hex_fill("EEF1F5")
    tgt_cell.alignment = Alignment(horizontal="center", vertical="center")
    tgt_cell.border    = _thin_border()
    ws.row_dimensions[tgt_row].height = 18

    for col_idx, d in enumerate(domains, start=3):
        col_letter = get_column_letter(col_idx)
        row_data = score_map.get(d)
        val = int(row_data["target"]) if row_data and row_data.get("target") is not None else None
        cell = ws[f"{col_letter}{tgt_row}"]
        cell.value     = val
        cell.font      = Font(bold=True, size=9, color="5A6570")
        cell.fill      = _hex_fill("EEF1F5")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

    # ── Row 12: Gap ───────────────────────────────────────────────────────────
    gap_row = 12
    gap_cell = ws[f"B{gap_row}"]
    gap_cell.value     = "Gap"
    gap_cell.font      = Font(bold=True, size=8)
    gap_cell.fill      = _hex_fill("EEF1F5")
    gap_cell.alignment = Alignment(horizontal="center", vertical="center")
    gap_cell.border    = _thin_border()
    ws.row_dimensions[gap_row].height = 18

    for col_idx, d in enumerate(domains, start=3):
        col_letter = get_column_letter(col_idx)
        row_data = score_map.get(d)
        val = round(float(row_data["gap"]), 1) if row_data and row_data.get("gap") is not None else None
        cell = ws[f"{col_letter}{gap_row}"]
        cell.value     = val
        cell.number_format = "0.0"
        # Colour gap cells: red for large gap, amber for moderate, green for small/no gap
        if val is not None:
            if val > 1.5:
                gap_fill = "F5C8C3"
            elif val > 0:
                gap_fill = "FDE9B2"
            else:
                gap_fill = "B7E2CD"
        else:
            gap_fill = "FFFFFF"
        cell.font      = Font(bold=True, size=9)
        cell.fill      = _hex_fill(gap_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

    # ── Legend at bottom ──────────────────────────────────────────────────────
    legend_row = 14
    ws[f"B{legend_row}"] = "Legend:"
    ws[f"B{legend_row}"].font = Font(bold=True, size=8)
    items = [
        ("B7E2CD", "Fully achieved (1.0)"),
        ("FDE9B2", "Partial progress (0–1)"),
        ("FFFFFF", "Not yet reached (0)"),
    ]
    for i, (hex_c, label) in enumerate(items):
        cl = get_column_letter(3 + i)
        cell = ws[f"{cl}{legend_row}"]
        cell.value     = label
        cell.font      = Font(size=8)
        cell.fill      = _hex_fill(hex_c)
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="center")

    # ── Copyright ─────────────────────────────────────────────────────────────
    ws[f"B{legend_row + 2}"] = "© Copyright 2018–2025 Hewlett Packard Enterprise Development LP"
    ws[f"B{legend_row + 2}"].font = Font(size=7, color="8A929A")

    # ── Freeze panes (keep level column and domain header visible) ────────────
    ws.freeze_panes = "C5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
