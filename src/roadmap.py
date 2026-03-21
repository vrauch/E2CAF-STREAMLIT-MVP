"""Roadmap rendering and Excel export for the gap-closure plan.

Given a roadmap JSON (from generate_roadmap_plan in ai_client.py), produces:
  - render_roadmap_gantt_html()  → Bootstrap HTML Gantt for st.components.v1.html()
  - generate_roadmap_excel()     → .xlsx bytes for st.download_button()
"""

from __future__ import annotations

import io
import math

from src.heatmap import DOMAIN_COLORS, _DEFAULT_COLOR

# ── Timeline unit configuration ──────────────────────────────────────────────

TIMELINE_UNITS = ["Sprints (2 wks)", "Weeks", "Quarters (13 wks)"]

_UNIT_WEEKS: dict[str, int] = {
    "Weeks": 1,
    "Sprints (2 wks)": 2,
    "Quarters (13 wks)": 13,
}
_UNIT_LABELS: dict[str, str] = {
    "Weeks": "W",
    "Sprints (2 wks)": "S",
    "Quarters (13 wks)": "Q",
}

# ── Priority colour palette ───────────────────────────────────────────────────

_PRIORITY_COLORS: dict[str, str] = {
    "Critical": "#DC2626",
    "High":     "#EA580C",
    "Medium":   "#D97706",
    "Low":      "#16A34A",
}
_PRIORITY_DEFAULT = "#6B7280"


def _week_to_period(week: int, timeline_unit: str) -> int:
    """Convert a 1-based week number to a 1-based period number."""
    pw = _UNIT_WEEKS.get(timeline_unit, 2)
    return math.ceil(week / pw)


def render_roadmap_gantt_html(roadmap: dict, timeline_unit: str = "Sprints (2 wks)") -> str:
    """Render a Bootstrap 5 Gantt HTML table from a roadmap JSON dict.

    Args:
        roadmap: The structured roadmap dict from generate_roadmap_plan.
        timeline_unit: One of TIMELINE_UNITS.

    Returns:
        Full HTML string suitable for st.components.v1.html().
    """
    total_weeks = int(roadmap.get("total_weeks", 24))
    period_size = _UNIT_WEEKS.get(timeline_unit, 2)
    total_periods = math.ceil(total_weeks / period_size)
    unit_short = _UNIT_LABELS.get(timeline_unit, "S")

    phases = roadmap.get("phases", [])
    critical_path = set(roadmap.get("critical_path", []))
    quick_wins = roadmap.get("quick_wins", [])

    # ── Column widths ─────────────────────────────────────────────────────────
    NAME_COL   = 220
    DOMAIN_COL = 130
    PRIO_COL   = 72
    PERIOD_COL = max(36, min(62, 780 // max(total_periods, 1)))
    total_width = NAME_COL + DOMAIN_COL + PRIO_COL + (PERIOD_COL * total_periods) + 24

    # ── Period header cells ───────────────────────────────────────────────────
    period_headers = "".join(
        f'<th style="min-width:{PERIOD_COL}px;max-width:{PERIOD_COL}px;'
        f'text-align:center;font-size:0.68rem;padding:3px 1px;white-space:nowrap;">'
        f'{unit_short}{p}</th>'
        for p in range(1, total_periods + 1)
    )

    # ── Table body ────────────────────────────────────────────────────────────
    rows_html = ""
    total_cols = 3 + total_periods  # name + domain + priority + periods

    for phase in phases:
        phase_name  = phase.get("name", "")
        story       = phase.get("story", "")
        description = phase.get("description", "")
        activities  = phase.get("activities", [])

        # Phase header row
        rows_html += (
            f'<tr style="background:#0F2744;color:#F9FAFB;">'
            f'<td colspan="{total_cols}" style="padding:7px 12px;font-weight:700;'
            f'font-size:0.88rem;letter-spacing:0.04em;">'
            f'<span style="opacity:0.5;font-size:0.7rem;margin-right:8px;'
            f'text-transform:uppercase;">Phase</span>{phase_name}'
            f'</td></tr>\n'
        )

        # Narrative row
        if story or description or activities:
            story_html = (
                f'<div style="font-style:italic;color:#6B7280;font-size:0.77rem;'
                f'margin-bottom:3px;">"{story}"</div>'
            ) if story else ""
            desc_html = (
                f'<div style="font-size:0.79rem;color:#374151;margin-bottom:3px;">'
                f'{description}</div>'
            ) if description else ""
            act_html = ""
            if activities:
                acts = "".join(f"<li>{a}</li>" for a in activities)
                act_html = (
                    f'<ul style="margin:3px 0 0;padding-left:16px;'
                    f'font-size:0.76rem;color:#374151;">{acts}</ul>'
                )
            rows_html += (
                f'<tr style="background:#F9FAFB;border-bottom:2px solid #D1D5DB;">'
                f'<td colspan="{total_cols}" style="padding:5px 12px 8px 22px;">'
                f'{story_html}{desc_html}{act_html}'
                f'</td></tr>\n'
            )

        # Initiative rows
        for init in phase.get("initiatives", []):
            init_name   = init.get("name", "")
            domain      = init.get("domain", "")
            priority    = init.get("priority", "")
            start_w     = int(init.get("start_week", 1))
            end_w       = int(init.get("end_week", total_weeks))
            outcome     = init.get("outcome", "")
            is_critical = init_name in critical_path

            domain_color = DOMAIN_COLORS.get(domain, _DEFAULT_COLOR)
            prio_color   = _PRIORITY_COLORS.get(priority, _PRIORITY_DEFAULT)

            # Convert weeks → periods and clamp to table bounds
            start_p = max(1, _week_to_period(start_w, timeline_unit))
            end_p   = min(total_periods, _week_to_period(end_w, timeline_unit))
            leading  = start_p - 1
            bar_span = max(1, end_p - start_p + 1)
            trailing = max(0, total_periods - end_p)

            period_cells = ""
            if leading > 0:
                period_cells += (
                    f'<td colspan="{leading}" '
                    f'style="background:#F9FAFB;border-color:#D1D5DB;padding:0;"></td>'
                )
            outcome_label = (
                f'<span style="font-size:0.67rem;color:#fff;padding:0 3px;'
                f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;display:block;">'
                f'{outcome}</span>'
            )
            period_cells += (
                f'<td colspan="{bar_span}" '
                f'style="background:{domain_color};padding:2px 3px;vertical-align:middle;" '
                f'title="{outcome}">{outcome_label}</td>'
            )
            if trailing > 0:
                period_cells += (
                    f'<td colspan="{trailing}" '
                    f'style="background:#F9FAFB;border-color:#D1D5DB;padding:0;"></td>'
                )

            star = " ⭐" if is_critical else ""
            rows_html += (
                f'<tr style="border-bottom:1px solid #D1D5DB;">'
                f'<td style="padding:4px 6px 4px 22px;font-size:0.79rem;font-weight:500;">'
                f'{init_name}{star}</td>'
                f'<td style="padding:4px;font-size:0.72rem;color:#6B7280;">{domain}</td>'
                f'<td style="padding:4px;text-align:center;">'
                f'<span style="background:{prio_color};color:#fff;border-radius:3px;'
                f'padding:1px 5px;font-size:0.68rem;">{priority}</span>'
                f'</td>'
                f'{period_cells}'
                f'</tr>\n'
            )

    # ── Quick wins section ────────────────────────────────────────────────────
    qw_html = ""
    if quick_wins:
        qw_items = "".join(
            f'<span style="background:#dcfce7;color:#15803d;border-radius:3px;'
            f'padding:2px 8px;margin:2px;display:inline-block;font-size:0.74rem;">'
            f'✓ {qw}</span>'
            for qw in quick_wins
        )
        qw_html = (
            f'<div style="margin-top:14px;padding:10px 14px;background:#f0fdf4;'
            f'border:1px solid #86efac;border-radius:6px;">'
            f'<div style="font-weight:600;font-size:0.84rem;color:#15803d;margin-bottom:5px;">'
            f'⚡ Quick Wins (completable within 2 weeks)</div>'
            f'<div>{qw_items}</div></div>'
        )

    return f"""<!DOCTYPE html>
<html><head>
  <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;600;700&display=swap" rel="stylesheet">
  <style>
    body {{ font-family:'Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; padding: 6px; margin: 0; background:#F9FAFB; }}
    .gantt-wrap {{ overflow-x: auto; }}
    .gantt-table {{ border-collapse: collapse; width: 100%; table-layout: fixed; min-width: {total_width}px; }}
    .gantt-table th, .gantt-table td {{ border: 1px solid #D1D5DB; vertical-align: middle; }}
    .gantt-table thead th {{ background: #0F2744; color: #F9FAFB; font-size: 0.76rem; padding: 5px 4px; }}
  </style>
</head><body>
  <div class="gantt-wrap">
    <table class="gantt-table">
      <thead>
        <tr>
          <th style="width:{NAME_COL}px;text-align:left;padding:6px 10px;">Initiative</th>
          <th style="width:{DOMAIN_COL}px;text-align:left;">Domain</th>
          <th style="width:{PRIO_COL}px;text-align:center;">Priority</th>
          {period_headers}
        </tr>
      </thead>
      <tbody>
        {rows_html}
      </tbody>
    </table>
  </div>
  {qw_html}
  <div style="margin-top:6px;font-size:0.68rem;color:#6B7280;">⭐ = Critical path initiative</div>
</body></html>"""


def generate_roadmap_excel(
    roadmap: dict,
    client_name: str = "",
    engagement_name: str = "",
    use_case_name: str = "",
) -> bytes:
    """Generate an Excel workbook from a roadmap dict.

    Returns bytes suitable for st.download_button().
    Sheets:
      1. Initiatives — flat table with all initiative details
      2. Phase Narratives — phase stories, descriptions, and activities
      3. Critical Path — critical path and quick wins
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Initiatives ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Initiatives"

    ws1.merge_cells("A1:J1")
    ws1["A1"] = f"Transformation Roadmap — {use_case_name}"
    ws1["A1"].font = Font(bold=True, size=14)

    ws1.merge_cells("A2:J2")
    ws1["A2"] = f"Client: {client_name}   |   Engagement: {engagement_name}"
    ws1["A2"].font = Font(italic=True, size=10, color="6B7280")

    headers1 = [
        "Phase", "Initiative", "Domain", "Priority",
        "Current Score", "Target", "Gap",
        "Start Week", "End Week", "Outcome",
    ]
    for col_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="0F2744")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row = 5
    for phase in roadmap.get("phases", []):
        phase_name = phase.get("name", "")
        for init in phase.get("initiatives", []):
            ws1.cell(row=row, column=1, value=phase_name)
            ws1.cell(row=row, column=2, value=init.get("name", ""))
            ws1.cell(row=row, column=3, value=init.get("domain", ""))
            ws1.cell(row=row, column=4, value=init.get("priority", ""))
            ws1.cell(row=row, column=5, value=init.get("current_score"))
            ws1.cell(row=row, column=6, value=init.get("target_score"))
            ws1.cell(row=row, column=7, value=init.get("gap"))
            ws1.cell(row=row, column=8, value=init.get("start_week"))
            ws1.cell(row=row, column=9, value=init.get("end_week"))
            ws1.cell(row=row, column=10, value=init.get("outcome", ""))
            row += 1

    for i, w in enumerate([18, 30, 22, 12, 14, 10, 8, 12, 10, 42], start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Phase Narratives ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Phase Narratives")

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Phase Narratives"
    ws2["A1"].font = Font(bold=True, size=13)

    headers2 = ["Phase", "Story", "Description", "Activities"]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="0F2744")

    row2 = 4
    for phase in roadmap.get("phases", []):
        ws2.cell(row=row2, column=1, value=phase.get("name", ""))
        ws2.cell(row=row2, column=2, value=phase.get("story", ""))
        ws2.cell(row=row2, column=3, value=phase.get("description", ""))
        acts = "\n".join(f"• {a}" for a in phase.get("activities", []))
        ws2.cell(row=row2, column=4, value=acts)
        for col in range(1, 5):
            ws2.cell(row=row2, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        row2 += 1

    for col, w in zip("ABCD", [20, 45, 45, 35]):
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: Critical Path & Quick Wins ───────────────────────────────────
    ws3 = wb.create_sheet("Critical Path")
    ws3["A1"] = "Critical Path"
    ws3["A1"].font = Font(bold=True, size=12)
    for i, name in enumerate(roadmap.get("critical_path", []), start=2):
        ws3.cell(row=i, column=1, value=f"• {name}")

    ws3["C1"] = "Quick Wins"
    ws3["C1"].font = Font(bold=True, size=12)
    for i, qw in enumerate(roadmap.get("quick_wins", []), start=2):
        ws3.cell(row=i, column=3, value=f"✓ {qw}")

    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["C"].width = 42

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
